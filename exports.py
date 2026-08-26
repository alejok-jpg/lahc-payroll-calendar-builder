import os
from datetime import date
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

WEEKDAYS_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

def format_date_with_weekday(d: Optional[date]) -> str:
    if not d:
        return "-"
    weekday_str = WEEKDAYS_ES[d.weekday()]
    return f"{weekday_str} {d.strftime('%d/%m/%Y')}"

def export_multiprocess_calendar_to_excel(
    events_by_process: Dict[str, List[Dict]],
    country: str,
    client_name: str = "CLIENTE",
    output_path: str = "Payroll_Calendar.xlsx",
    logo_path: Optional[str] = None
) -> str:
    wb = openpyxl.Workbook()

    font_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    
    fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_master_hdr = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # -------------------------------------------------------------
    # 1. PESTAÑA CONSOLIDADA: MASTER_VIEW
    # -------------------------------------------------------------
    ws_master = wb.active
    ws_master.title = "MASTER_VIEW"
    ws_master.views.sheetView[0].showGridLines = True

    all_events: List[Dict] = []
    for proc_name, ev_list in events_by_process.items():
        all_events.extend(ev_list)

    all_events.sort(key=lambda x: (x["date"], x.get("time", "00:00"), x["process"]))

    start_row_m = 1
    has_logo = logo_path and os.path.exists(logo_path)
    if has_logo:
        img_m = Image(logo_path)
        img_m.width = 150
        img_m.height = 50
        ws_master.add_image(img_m, "A1")
        ws_master.row_dimensions[1].height = 22
        ws_master.row_dimensions[2].height = 22
        ws_master.row_dimensions[3].height = 12
        start_row_m = 5

    ws_master.merge_cells(start_row=start_row_m, start_column=1, end_row=start_row_m, end_column=8)
    m_title = ws_master.cell(
        row=start_row_m,
        column=1,
        value=f"PAYROLL MASTER SCHEDULE - {client_name.upper()} ({country.upper()})"
    )
    m_title.font = font_title
    m_title.fill = fill_title
    m_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_master.row_dimensions[start_row_m].height = 30

    m_hdr_row = start_row_m + 1
    m_headers = ["Fecha Operativa", "Proceso", "Periodo / Ciclo", "Actividad SLA", "Owner", "Offset (BH)", "Horario (SLA)", "Notas / Control"]
    ws_master.row_dimensions[m_hdr_row].height = 24

    for col_num, h_text in enumerate(m_headers, 1):
        cell = ws_master.cell(row=m_hdr_row, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = fill_master_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    m_curr_row = m_hdr_row + 1
    for row_idx, ev in enumerate(all_events):
        ws_master.row_dimensions[m_curr_row].height = 20
        use_zebra = (row_idx % 2 == 1)

        c_date  = ws_master.cell(row=m_curr_row, column=1, value=format_date_with_weekday(ev["date"]))
        c_proc  = ws_master.cell(row=m_curr_row, column=2, value=ev["process"])
        c_per   = ws_master.cell(row=m_curr_row, column=3, value=ev["period"])
        c_act   = ws_master.cell(row=m_curr_row, column=4, value=ev["activity"])
        c_owner = ws_master.cell(row=m_curr_row, column=5, value=ev.get("owner", "ADP"))
        c_off   = ws_master.cell(row=m_curr_row, column=6, value=f"{ev.get('offset_bh', 0)} BH")
        c_time  = ws_master.cell(row=m_curr_row, column=7, value=ev.get("time", "18:00"))
        c_note  = ws_master.cell(row=m_curr_row, column=8, value="")

        c_date.alignment  = Alignment(horizontal="center", vertical="center")
        c_proc.alignment  = Alignment(horizontal="center", vertical="center")
        c_per.alignment   = Alignment(horizontal="center", vertical="center")
        c_owner.alignment = Alignment(horizontal="center", vertical="center")
        c_off.alignment   = Alignment(horizontal="center", vertical="center")
        c_time.alignment  = Alignment(horizontal="center", vertical="center")

        for c in (c_date, c_proc, c_per, c_act, c_owner, c_off, c_time, c_note):
            c.font = font_body
            c.border = thin_border
            if use_zebra:
                c.fill = fill_zebra

        m_curr_row += 1

    for col in ws_master.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_master.column_dimensions[col_letter].width = max(max_len + 4, 16)

    # -------------------------------------------------------------
    # 2. PESTAÑAS INDIVIDUALES POR PROCESO
    # -------------------------------------------------------------
    for proc_name, events in events_by_process.items():
        if not events:
            continue
        
        ws = wb.create_sheet(title=proc_name[:31])
        ws.views.sheetView[0].showGridLines = True

        period_order = []
        for ev in sorted(events, key=lambda x: x["pay_date"]):
            if ev["period"] not in period_order:
                period_order.append(ev["period"])

        total_cols = 5 + len(period_order)  # Process, Activity, Owner, Offset, Time + Periodos

        start_row = 1
        if has_logo:
            img = Image(logo_path)
            img.width = 150
            img.height = 50
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 22
            ws.row_dimensions[3].height = 12
            start_row = 5

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
        title_cell = ws.cell(
            row=start_row,
            column=1,
            value=f"PAYROLL CALENDAR - {client_name.upper()} | {proc_name} ({country.upper()})"
        )
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row].height = 30

        header_row = start_row + 1
        headers = ["Process", "Activity", "Owner", "Offset (BH)", "Time (SLA)"] + period_order
        ws.row_dimensions[header_row].height = 24

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Matriz: (Process, Activity, Owner, Offset, Time)
        matrix: Dict[tuple, Dict[str, date]] = {}
        for ev in events:
            key = (ev["process"], ev["activity"], ev.get("owner", "ADP"), ev["offset_bh"], ev.get("time", "18:00"))
            if key not in matrix:
                matrix[key] = {}
            matrix[key][ev["period"]] = ev["date"]

        current_row = header_row + 1
        for row_idx, ((proc, act, owner_val, offset, time_val), dates_by_period) in enumerate(matrix.items()):
            ws.row_dimensions[current_row].height = 20
            use_zebra = (row_idx % 2 == 1)

            c1 = ws.cell(row=current_row, column=1, value=proc)
            c2 = ws.cell(row=current_row, column=2, value=act)
            c3 = ws.cell(row=current_row, column=3, value=owner_val)
            c4 = ws.cell(row=current_row, column=4, value=f"{offset} BH")
            c5 = ws.cell(row=current_row, column=5, value=time_val)

            c3.alignment = Alignment(horizontal="center", vertical="center")
            c4.alignment = Alignment(horizontal="center", vertical="center")
            c5.alignment = Alignment(horizontal="center", vertical="center")

            for c in (c1, c2, c3, c4, c5):
                c.font = font_body
                c.border = thin_border
                if use_zebra:
                    c.fill = fill_zebra

            for col_offset, period in enumerate(period_order):
                col_num = 6 + col_offset
                date_val = dates_by_period.get(period)
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = format_date_with_weekday(date_val)
                cell.font = font_body
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                if use_zebra:
                    cell.fill = fill_zebra

            current_row += 1

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(output_path)
    return output_path