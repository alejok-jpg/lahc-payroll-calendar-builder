from datetime import date, timedelta
from typing import Dict, List
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from processes import ProcessType

def generate_multiprocess_template(output_path: str = "Template_PayDates.xlsx", year: int = 2027) -> str:
    """
    Genera un archivo Excel con una pestaña por cada proceso disponible.
    Para TERMINATION crea una tabla de configuración por regla semanal y rango de fechas.
    """
    wb = openpyxl.Workbook()
    
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    font_body_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_term_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_label = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for proc in ProcessType:
        ws = wb.create_sheet(title=proc.name)
        ws.views.sheetView[0].showGridLines = True

        # Pestaña especial para Bajas (Configuración Semanal)
        if proc == ProcessType.TERMINATION:
            # Encabezados
            headers = ["Configuración de Bajas", "Valor", "Instrucciones / Ejemplos"]
            for col_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h_text)
                cell.font = font_header
                cell.fill = fill_term_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            ws.row_dimensions[1].height = 25

            # Parámetros
            config_rows = [
                ("Días de Corte Semanales", "Martes, Jueves", "Ej: Martes, Jueves (o Lunes, Miércoles, Viernes)"),
                ("Fecha Inicio (DD/MM/YYYY)", f"01/01/{year}", "Inicio del periodo a proyectar"),
                ("Fecha Fin (DD/MM/YYYY)", f"31/12/{year}", "Fin del periodo a proyectar")
            ]

            for row_idx, (label, val, guide) in enumerate(config_rows, 2):
                c1 = ws.cell(row=row_idx, column=1, value=label)
                c2 = ws.cell(row=row_idx, column=2, value=val)
                c3 = ws.cell(row=row_idx, column=3, value=guide)

                c1.font = font_body_bold
                c1.fill = fill_label
                c2.font = font_body
                c3.font = font_body

                c1.border = thin_border
                c2.border = thin_border
                c3.border = thin_border

                c2.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[row_idx].height = 22

            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 45

        # Pestañas estándar basadas en Pay Day
        else:
            if proc == ProcessType.BIWEEKLY:
                header_col = "PAY_DATE (DD/MM/YYYY)"
                sample_dates = [
                    date(year, 1, 15), date(year, 1, 29),
                    date(year, 2, 15), date(year, 2, 26)
                ]
            elif proc == ProcessType.MONTHLY:
                header_col = "PAY_DATE (DD/MM/YYYY)"
                sample_dates = [date(year, m, 28) for m in range(1, 13)]
            else:
                header_col = "PAY_DATE (DD/MM/YYYY)"
                sample_dates = [date(year, 6, 30), date(year, 12, 20)]

            cell_hdr = ws.cell(row=1, column=1, value=header_col)
            cell_hdr.font = font_header
            cell_hdr.fill = fill_header
            cell_hdr.alignment = Alignment(horizontal="center", vertical="center")
            cell_hdr.border = thin_border
            ws.row_dimensions[1].height = 24

            for row_idx, d_val in enumerate(sample_dates, 2):
                cell_data = ws.cell(row=row_idx, column=1, value=d_val.strftime("%d/%m/%Y"))
                cell_data.font = font_body
                cell_data.alignment = Alignment(horizontal="center", vertical="center")
                cell_data.border = thin_border
                ws.row_dimensions[row_idx].height = 20

            ws.column_dimensions["A"].width = 30

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(output_path)
    return output_path


def _parse_weekdays(text: str) -> List[int]:
    """Interpreta días de la semana en español/inglés y devuelve lista de enteros (0=Lun..6=Dom)."""
    mapping = {
        'lun': 0, 'mon': 0, 'lunes': 0, 'monday': 0,
        'mar': 1, 'tue': 1, 'martes': 1, 'tuesday': 1,
        'mie': 2, 'mié': 2, 'wed': 2, 'miercoles': 2, 'miércoles': 2, 'wednesday': 2,
        'jue': 3, 'thu': 3, 'jueves': 3, 'thursday': 3,
        'vie': 4, 'fri': 4, 'viernes': 4, 'friday': 4,
        'sab': 5, 'sáb': 5, 'sat': 5, 'sabado': 5, 'sábado': 5, 'saturday': 5,
        'dom': 6, 'sun': 6, 'domingo': 6, 'sunday': 6
    }
    found = set()
    for word in str(text).lower().replace(',', ' ').replace(';', ' ').replace('/', ' ').split():
        for k, v in mapping.items():
            if word.startswith(k):
                found.add(v)
                break
    return sorted(list(found))


def load_multiprocess_pay_dates_from_excel(file_content) -> Dict[str, List[date]]:
    """
    Lee el archivo Excel y extrae las fechas por proceso.
    Para TERMINATION soporta tanto la tabla de configuración semanal como una lista de fechas tradicional.
    """
    excel_file = pd.ExcelFile(file_content)
    result: Dict[str, List[date]] = {}

    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        if df.empty:
            continue

        # Tratamiento especial para TERMINATION
        if sheet_name == ProcessType.TERMINATION.name:
            # Caso 1: Formato de Configuración Semanal (Opción A)
            if len(df.columns) >= 2 and any("Días" in str(x) or "Dias" in str(x) or "Configuración" in str(x) for x in df.iloc[:, 0].dropna()):
                try:
                    cfg_dict = {}
                    for _, row in df.iterrows():
                        key_str = str(row.iloc[0]).lower()
                        val = row.iloc[1]
                        if "día" in key_str or "dia" in key_str:
                            cfg_dict["weekdays"] = _parse_weekdays(str(val))
                        elif "inicio" in key_str:
                            cfg_dict["start"] = pd.to_datetime(val, dayfirst=True).date()
                        elif "fin" in key_str:
                            cfg_dict["end"] = pd.to_datetime(val, dayfirst=True).date()

                    weekdays = cfg_dict.get("weekdays", [1, 3])  # Default Martes/Jueves
                    start_d = cfg_dict.get("start", date(2027, 1, 1))
                    end_d = cfg_dict.get("end", date(2027, 12, 31))

                    generated_cutoffs = []
                    curr = start_d
                    while curr <= end_d:
                        if curr.weekday() in weekdays:
                            generated_cutoffs.append(curr)
                        curr += timedelta(days=1)

                    if generated_cutoffs:
                        result[sheet_name] = generated_cutoffs
                        continue
                except Exception:
                    pass

        # Caso 2: Formato estándar de columna de fechas
        col_name = None
        for c in df.columns:
            clean_c = str(c).upper()
            if any(k in clean_c for k in ["PAY_DATE", "TERMINATION", "CUT_OFF", "FECHA", "DATE"]):
                col_name = c
                break

        if not col_name:
            col_name = df.columns[0]

        extracted_dates = []
        for val in df[col_name].dropna():
            try:
                dt = pd.to_datetime(val, dayfirst=True)
                extracted_dates.append(dt.date())
            except Exception:
                continue

        if extracted_dates:
            result[sheet_name] = sorted(list(set(extracted_dates)))

    return result